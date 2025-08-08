import PyPDF2
import xml.etree.ElementTree as ET
import os
import shutil
from rich.console import Console
from rich.progress import Progress
import zipfile
import warnings
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

try:
    import rarfile
except ImportError:
    rarfile = None


console = Console()


def leer_pdf(ruta_pdf, password=None):
    # Suprime warnings de PyPDF2 (incluyendo "incorrect startxref pointer")
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        with open(ruta_pdf, "rb") as archivo:
            lector = PyPDF2.PdfReader(archivo)
            if lector.is_encrypted:
                if password:
                    try:
                        lector.decrypt(password)
                    except Exception:
                        raise Exception(
                            "No se pudo desencriptar el PDF con la contraseña encontrada."
                        )
                else:
                    raise Exception("PDF encriptado y no se proporcionó contraseña.")
            texto = ""
            for pagina in lector.pages:
                texto += pagina.extract_text()
            return texto


def obtener_nombre_receptor(xml_factura):
    tree = ET.parse(xml_factura)
    root = tree.getroot()
    for elem in root.iter():
        if elem.tag.endswith("Receptor"):
            for subelem in elem.iter():
                if subelem.tag.endswith("Nombre"):
                    return subelem.text
    return "Desconocido"


def obtener_cedula_receptor(xml_factura):
    tree = ET.parse(xml_factura)
    root = tree.getroot()
    # Si es MensajeHacienda o MensajeReceptor, buscar NumeroCedulaReceptor directamente bajo el nodo raíz
    if root.tag.endswith("MensajeHacienda") or root.tag.endswith("MensajeReceptor"):
        for elem in root.iter():
            if elem.tag.endswith("NumeroCedulaReceptor"):
                return elem.text
    # Si es factura, nota o tiquete, buscar en Receptor
    for elem in root.iter():
        if elem.tag.endswith("Receptor"):
            for subelem in elem.iter():
                if subelem.tag.endswith("Numero"):
                    return subelem.text
    return "Desconocido"


def descomprimir_archivos_en_carpeta(carpeta):
    for root, dirs, files in os.walk(carpeta):
        for file in files:
            ruta = os.path.join(root, file)
            if file.lower().endswith(".zip"):
                try:
                    with zipfile.ZipFile(ruta, "r") as zf:
                        zf.extractall(root)
                except Exception as e:
                    print(f"Error descomprimiendo {ruta}: {e}")
            elif file.lower().endswith(".rar") and rarfile:
                try:
                    with rarfile.RarFile(ruta) as rf:
                        rf.extractall(root)
                except Exception as e:
                    print(f"Error descomprimiendo {ruta}: {e}")


def obtener_todos_los_archivos(carpeta):
    archivos = []
    for root, dirs, files in os.walk(carpeta):
        for file in files:
            archivos.append(os.path.join(root, file))
    return archivos


def limpiar_xml(filepath):
    # Detecta el cierre de cualquier etiqueta raíz conocida y elimina basura posterior
    # Además, evita reescribir si no hay basura
    cierres = [
        "</MensajeHacienda>",
        "</MensajeReceptor>",
        "</FacturaElectronica>",
        "</NotaCreditoElectronica>",
        "</TiqueteElectronico>",
    ]
    with open(filepath, "r", encoding="utf-8") as f:
        contenido = f.read()
    for cierre in cierres:
        idx = contenido.find(cierre)
        if idx != -1 and idx + len(cierre) < len(contenido):
            contenido = contenido[: idx + len(cierre)]
            with open(filepath, "w", encoding="utf-8") as f:
                f.write(contenido)
            break
    # Si no hay basura, no reescribe el archivo


def extraer_datos_factura(xml_path):
    tree = ET.parse(xml_path)
    root = tree.getroot()
    datos = {}
    # Clave, Fecha, Emisor, Cedula Emisor, Receptor, Cedula Receptor, Subtotal, Total, Impuesto, TipoMoneda, TipoCambio
    for elem in root.iter():
        if elem.tag.endswith("Clave"):
            datos["Clave"] = elem.text
        if elem.tag.endswith("FechaEmision"):
            datos["Fecha"] = elem.text
        if elem.tag.endswith("Emisor"):
            for subelem in elem.iter():
                if subelem.tag.endswith("Nombre"):
                    datos["Emisor"] = subelem.text
                if subelem.tag.endswith("Numero"):
                    datos["CedulaEmisor"] = subelem.text
        if elem.tag.endswith("Receptor"):
            for subelem in elem.iter():
                if subelem.tag.endswith("Nombre"):
                    datos["Receptor"] = subelem.text
                if subelem.tag.endswith("Numero"):
                    datos["CedulaReceptor"] = subelem.text
        if elem.tag.endswith("TotalComprobante"):
            datos["Total"] = elem.text
        if elem.tag.endswith("TotalImpuesto"):
            datos["Impuesto"] = elem.text
        if elem.tag.endswith("TotalVenta"):
            datos["Subtotal"] = elem.text
        if elem.tag.endswith("CodigoTipoMoneda"):
            for subelem in elem.iter():
                if subelem.tag.endswith("CodigoMoneda"):
                    datos["TipoMoneda"] = subelem.text
                if subelem.tag.endswith("TipoCambio"):
                    datos["TipoCambio"] = subelem.text
    # Si no se encuentra, dejar vacío
    datos.setdefault("Subtotal", "")
    datos.setdefault("TipoMoneda", "")
    datos.setdefault("TipoCambio", "")
    return datos


def extraer_datos_tiquete(xml_path):
    # Igual que factura, pero para TiqueteElectronico
    return extraer_datos_factura(xml_path)


def extraer_datos_nota_credito(xml_path):
    tree = ET.parse(xml_path)
    root = tree.getroot()
    datos = {}
    for elem in root.iter():
        if elem.tag.endswith("Clave"):
            datos["Clave"] = elem.text
        if elem.tag.endswith("FechaEmision"):
            datos["Fecha"] = elem.text
        if elem.tag.endswith("Emisor"):
            for subelem in elem.iter():
                if subelem.tag.endswith("Nombre"):
                    datos["Emisor"] = subelem.text
                if subelem.tag.endswith("Numero"):
                    datos["CedulaEmisor"] = subelem.text
        if elem.tag.endswith("Receptor"):
            for subelem in elem.iter():
                if subelem.tag.endswith("Nombre"):
                    datos["Receptor"] = subelem.text
                if subelem.tag.endswith("Numero"):
                    datos["CedulaReceptor"] = subelem.text
        if elem.tag.endswith("TotalComprobante"):
            datos["Total"] = elem.text
        if elem.tag.endswith("TotalImpuesto"):
            datos["Impuesto"] = elem.text
        if elem.tag.endswith("TotalVenta"):
            datos["Subtotal"] = elem.text
        if elem.tag.endswith("CodigoTipoMoneda"):
            for subelem in elem.iter():
                if subelem.tag.endswith("CodigoMoneda"):
                    datos["TipoMoneda"] = subelem.text
                if subelem.tag.endswith("TipoCambio"):
                    datos["TipoCambio"] = subelem.text
        if elem.tag.endswith("InformacionReferencia"):
            for subelem in elem.iter():
                if subelem.tag.endswith("Numero"):
                    datos["Referencia"] = subelem.text
    datos.setdefault("Subtotal", "")
    datos.setdefault("TipoMoneda", "")
    datos.setdefault("TipoCambio", "")
    return datos


def obtener_mes(fecha_str):
    # fecha_str: "2022-06-10T15:19:52-06:00"
    try:
        fecha = datetime.strptime(fecha_str[:19], "%Y-%m-%dT%H:%M:%S")
        return fecha.strftime("%Y-%m")
    except Exception:
        return "SinFecha"


def obtener_mes_anio(fecha_str):
    # fecha_str: "2022-06-10T15:19:52-06:00"
    try:
        fecha = datetime.strptime(fecha_str[:19], "%Y-%m-%dT%H:%M:%S")
        return fecha.strftime("%m"), fecha.strftime("%Y")
    except Exception:
        return "SinMes", "SinAnio"


def obtener_mes_anio_desde_clave(clave):
    # clave: 506DDMMYY... (primeros 9 dígitos), DD=dia, MM=mes, YY=año (2 dígitos)
    try:
        if clave and len(clave) >= 9 and clave.startswith("506"):
            dia = clave[3:5]
            mes = clave[5:7]
            anio = clave[7:9]
            # Validación mejorada
            if int(mes) >= 1 and int(mes) <= 12 and int(dia) >= 1 and int(dia) <= 31:
                # Convierte año de dos dígitos a cuatro dígitos (asume 2000+)
                anio_full = f"20{anio}"
                return mes, anio_full
    except Exception:
        pass
    return "SinMes", "SinAnio"


def generar_reportes_excel(facturas, tiquetes, notas_credito):
    agrupado = {}
    for f in facturas + tiquetes:
        datos = extraer_datos_factura(f["path"])
        cedula = datos.get("CedulaReceptor", "Desconocido")
        mes = f.get("mes", "SinMes")
        anio = f.get("anio", "SinAnio")
        agrupado.setdefault(cedula, {}).setdefault(anio, {}).setdefault(
            mes, {"facturas": [], "notas": []}
        )
        agrupado[cedula][anio][mes]["facturas"].append(datos)
    for nc in notas_credito:
        datos = extraer_datos_nota_credito(nc["path"])
        cedula = datos.get("CedulaReceptor", "Desconocido")
        mes = nc.get("mes", "SinMes")
        anio = nc.get("anio", "SinAnio")
        agrupado.setdefault(cedula, {}).setdefault(anio, {}).setdefault(
            mes, {"facturas": [], "notas": []}
        )
        agrupado[cedula][anio][mes]["notas"].append(datos)

    for cedula, anios in agrupado.items():
        for anio, meses in anios.items():
            cedula_anio_dir = os.path.join("procesados", cedula, anio)
            os.makedirs(cedula_anio_dir, exist_ok=True)
            for mes, datos_mes in meses.items():
                wb = openpyxl.Workbook()
                if wb.active is not None:
                    wb.remove(wb.active)
                ws_f = wb.create_sheet("Facturas")
                columnas_f = [
                    "Clave",
                    "Fecha",
                    "Emisor",
                    "CedulaEmisor",
                    "Receptor",
                    "CedulaReceptor",
                    "Subtotal",
                    "Total",
                    "Impuesto",
                    "TipoMoneda",
                    "TipoCambio",
                ]
                ws_f.append(columnas_f)
                for fila in datos_mes["facturas"]:
                    fecha = fila.get("Fecha", "")
                    try:
                        fecha_dt = datetime.strptime(fecha[:10], "%Y-%m-%d")
                        fecha_fmt = fecha_dt.strftime("%d/%m/%Y")
                    except Exception:
                        fecha_fmt = fecha
                    moneda = fila.get("TipoMoneda", "")
                    tipo_cambio = fila.get("TipoCambio", "")
                    try:
                        tc = float(tipo_cambio) if tipo_cambio else 1.0
                    except Exception:
                        tc = 1.0

                    def convertir(valor):
                        try:
                            v = float(valor)
                            if moneda and moneda != "CRC" and tc > 0:
                                return round(v * tc, 2)
                            return round(v, 2)
                        except Exception:
                            return valor

                    subtotal = convertir(fila.get("Subtotal", ""))
                    total = convertir(fila.get("Total", ""))
                    impuesto = convertir(fila.get("Impuesto", ""))
                    ws_f.append(
                        [
                            fila.get("Clave", ""),
                            fecha_fmt,
                            fila.get("Emisor", ""),
                            fila.get("CedulaEmisor", ""),
                            fila.get("Receptor", ""),
                            fila.get("CedulaReceptor", ""),
                            subtotal,
                            total,
                            impuesto,
                            moneda,
                            tipo_cambio,
                        ]
                    )
                for col in ws_f.columns:
                    max_length = max(
                        len(str(cell.value)) if cell.value is not None else 0
                        for cell in col
                    )
                    ws_f.column_dimensions[get_column_letter(col[0].column)].width = (
                        max_length + 2
                    )
                ws_nc = wb.create_sheet("NotasCredito")
                columnas_nc = [
                    "Clave",
                    "Fecha",
                    "Emisor",
                    "CedulaEmisor",
                    "Receptor",
                    "CedulaReceptor",
                    "Subtotal",
                    "Total",
                    "Impuesto",
                    "TipoMoneda",
                    "TipoCambio",
                    "Referencia",
                ]
                ws_nc.append(columnas_nc)
                for fila in datos_mes["notas"]:
                    fecha = fila.get("Fecha", "")
                    try:
                        fecha_dt = datetime.strptime(fecha[:10], "%Y-%m-%d")
                        fecha_fmt = fecha_dt.strftime("%d/%m/%Y")
                    except Exception:
                        fecha_fmt = fecha
                    moneda = fila.get("TipoMoneda", "")
                    tipo_cambio = fila.get("TipoCambio", "")
                    try:
                        tc = float(tipo_cambio) if tipo_cambio else 1.0
                    except Exception:
                        tc = 1.0
                    subtotal = convertir(fila.get("Subtotal", ""))
                    total = convertir(fila.get("Total", ""))
                    impuesto = convertir(fila.get("Impuesto", ""))
                    ws_nc.append(
                        [
                            fila.get("Clave", ""),
                            fecha_fmt,
                            fila.get("Emisor", ""),
                            fila.get("CedulaEmisor", ""),
                            fila.get("Receptor", ""),
                            fila.get("CedulaReceptor", ""),
                            subtotal,
                            total,
                            impuesto,
                            moneda,
                            tipo_cambio,
                            fila.get("Referencia", ""),
                        ]
                    )
                for col in ws_nc.columns:
                    max_length = max(
                        len(str(cell.value)) if cell.value is not None else 0
                        for cell in col
                    )
                    ws_nc.column_dimensions[get_column_letter(col[0].column)].width = (
                        max_length + 2
                    )
                # Cambia el nombre del archivo a MES-ANIO.xlsx
                ruta_libro = os.path.join(cedula_anio_dir, f"{mes}-{anio}.xlsx")
                wb.save(ruta_libro)


def main():
    para_procesar = "examples"
    base_dir = os.path.dirname(__file__)
    procesados_dir = os.path.join(base_dir, "procesados")
    otros_dir = os.path.join(procesados_dir, "Otros")
    os.makedirs(procesados_dir, exist_ok=True)
    os.makedirs(otros_dir, exist_ok=True)
    ruta_carpeta = os.path.join(base_dir, para_procesar)
    descomprimir_archivos_en_carpeta(ruta_carpeta)
    archivos = obtener_todos_los_archivos(ruta_carpeta)

    facturas = []
    notas_credito = []
    mensajes_hacienda = []
    tiquetes = []
    pdfs = []
    otros = []
    mensajes_receptor = []

    def nombre_corto(nombre, maxlen=30):
        nombre = os.path.basename(nombre)
        return (nombre[:maxlen] + "...") if len(nombre) > maxlen else nombre

    # Barra de progreso para clasificación
    with Progress(transient=True, expand=True) as progress:
        task = progress.add_task(
            "[cyan]Clasificando archivos... 0/{}".format(len(archivos)),
            total=len(archivos),
        )
        contador = 0
        # Clasificación
        for idx, archivo in enumerate(archivos):
            if archivo.lower().endswith(".xml"):
                limpiar_xml(archivo)  # <-- limpiar antes de parsear
                try:
                    tree = ET.parse(archivo)
                    root = tree.getroot()
                    if root.tag.endswith("FacturaElectronica"):
                        clave = None
                        for elem in root.iter():
                            if elem.tag.endswith("Clave"):
                                clave = elem.text
                                break
                        facturas.append({"clave": clave, "path": archivo})
                    elif root.tag.endswith("NotaCreditoElectronica"):
                        clave = None
                        cedula = None
                        referencia = None
                        for elem in root.iter():
                            if elem.tag.endswith("Clave"):
                                clave = elem.text
                            if elem.tag.endswith("Emisor"):
                                for subelem in elem.iter():
                                    if subelem.tag.endswith("Numero"):
                                        cedula = subelem.text
                            if elem.tag.endswith("InformacionReferencia"):
                                for subelem in elem.iter():
                                    if subelem.tag.endswith("Numero"):
                                        referencia = subelem.text
                        notas_credito.append(
                            {
                                "clave": clave,
                                "path": archivo,
                                "cedula": cedula,
                                "referencia": referencia,
                            }
                        )
                    elif root.tag.endswith("MensajeHacienda"):
                        clave = None
                        cedula = None
                        for elem in root.iter():
                            if elem.tag.endswith("Clave"):
                                clave = elem.text
                            if elem.tag.endswith("NumeroCedulaReceptor"):
                                cedula = elem.text
                        mensajes_hacienda.append(
                            {"clave": clave, "path": archivo, "cedula": cedula}
                        )
                    elif root.tag.endswith("TiqueteElectronico"):
                        clave = None
                        for elem in root.iter():
                            if elem.tag.endswith("Clave"):
                                clave = elem.text
                                break
                        tiquetes.append({"clave": clave, "path": archivo})
                    elif root.tag.endswith("MensajeReceptor"):
                        clave = None
                        cedula = None
                        for elem in root.iter():
                            if elem.tag.endswith("Clave"):
                                clave = elem.text
                            if elem.tag.endswith("NumeroCedulaReceptor"):
                                cedula = elem.text
                        mensajes_receptor.append(
                            {"clave": clave, "path": archivo, "cedula": cedula}
                        )
                    else:
                        otros.append(archivo)
                except Exception:
                    otros.append(archivo)
            elif archivo.lower().endswith(".pdf"):
                try:
                    texto = leer_pdf(archivo)
                    import re

                    # Unir el texto para buscar la clave
                    texto_unido = texto.replace("\n", "").replace("\r", "")
                    # Solo acepta claves que sean exactamente 50 dígitos (no números de factura, ni secuencias cortas/largas)
                    match = re.search(r"(?<!\d)(\d{50})(?!\d)", texto_unido)
                    clave = match.group(1) if match else None
                    # Solo agregar si la clave cumple exactamente con el patrón de 50 dígitos
                    if clave:
                        pdfs.append({"clave": clave, "path": archivo})
                    else:
                        otros.append(archivo)
                except Exception:
                    otros.append(archivo)
            else:
                otros.append(archivo)
            progress.update(
                task,
                advance=1,
                description=f"[cyan]Clasificando archivos... {idx+1}/{len(archivos)} | {nombre_corto(archivo)}",
            )

    # Organización
    # 1. Facturas
    for factura in facturas:
        clave = factura["clave"]
        path = factura["path"]
        cedula = obtener_cedula_receptor(path)  # <-- ya usa receptor
        if not clave or not cedula:
            otros.append(path)
            continue
        cedula_dir = os.path.join(procesados_dir, cedula)
        facturas_dir = os.path.join(cedula_dir, "Facturas")
        os.makedirs(facturas_dir, exist_ok=True)
        destino = os.path.join(facturas_dir, f"FE-{clave}.xml")
        shutil.copy2(path, destino)

    # 2. Notas de crédito
    for nc in notas_credito:
        clave = nc["clave"]
        path = nc["path"]
        # Cambia para usar la cédula del receptor
        cedula = obtener_cedula_receptor(path) or "Desconocido"
        if not clave or not cedula:
            otros.append(path)
            continue
        cedula_dir = os.path.join(procesados_dir, cedula)
        nc_dir = os.path.join(cedula_dir, "NotasCredito")
        os.makedirs(nc_dir, exist_ok=True)
        destino = os.path.join(nc_dir, f"NC-{clave}.xml")
        shutil.copy2(path, destino)

    # 3. Mensajes Hacienda
    for mh in mensajes_hacienda:
        clave = mh["clave"]
        path = mh["path"]
        cedula = obtener_cedula_receptor(path) or "Desconocido"
        if not clave or not cedula:
            otros.append(path)
            continue
        cedula_dir = os.path.join(procesados_dir, cedula)
        respuestas_dir = os.path.join(cedula_dir, "Respuestas")
        os.makedirs(respuestas_dir, exist_ok=True)
        destino = os.path.join(respuestas_dir, f"MH-{clave}.xml")
        shutil.copy2(path, destino)

    # 3b. Mensajes Receptor
    for mr in mensajes_receptor:
        clave = mr["clave"]
        path = mr["path"]
        cedula = obtener_cedula_receptor(path) or "Desconocido"
        if not clave or not cedula:
            otros.append(path)
            continue
        cedula_dir = os.path.join(procesados_dir, cedula)
        respuestas_dir = os.path.join(cedula_dir, "Respuestas")
        os.makedirs(respuestas_dir, exist_ok=True)
        destino = os.path.join(respuestas_dir, f"MR-{clave}.xml")
        shutil.copy2(path, destino)

    # 4. Tiquetes
    for tiquete in tiquetes:
        clave = tiquete["clave"]
        path = tiquete["path"]
        cedula = obtener_cedula_receptor(path)
        if not clave or not cedula:
            otros.append(path)
            continue
        cedula_dir = os.path.join(procesados_dir, cedula)
        tiquetes_dir = os.path.join(cedula_dir, "Tiquetes")
        os.makedirs(tiquetes_dir, exist_ok=True)
        destino = os.path.join(tiquetes_dir, f"TQ-{clave}.xml")
        shutil.copy2(path, destino)

    # 5. PDFs
    for pdf in pdfs:
        clave = pdf["clave"]
        path = pdf["path"]
        if not clave:
            otros.append(path)
            continue
        # Buscar la factura asociada por clave
        factura = next((f for f in facturas if f["clave"] == clave), None)
        cedula = None
        if factura:
            cedula = obtener_cedula_receptor(factura["path"])
        else:
            # Buscar nota de crédito asociada
            nc = next((n for n in notas_credito if n["clave"] == clave), None)
            cedula = obtener_cedula_receptor(nc["path"]) if nc else "Desconocido"
        if not cedula:
            cedula = "Desconocido"
        cedula_dir = os.path.join(procesados_dir, cedula)
        pdf_dir = os.path.join(cedula_dir, "PDF")
        os.makedirs(pdf_dir, exist_ok=True)
        destino = os.path.join(pdf_dir, f"PDF-{clave}.pdf")
        shutil.copy2(path, destino)

    # 6. Otros
    for archivo in otros:
        ext = os.path.splitext(archivo)[1].lower().replace(".", "") or "sin_extension"
        ext_dir = os.path.join(otros_dir, ext)
        os.makedirs(ext_dir, exist_ok=True)
        try:
            shutil.copy2(archivo, os.path.join(ext_dir, os.path.basename(archivo)))
        except Exception:
            pass

    # Generar reportes Excel por receptor
    generar_reportes_excel(facturas, tiquetes, notas_credito)


def main_procesar_facturas(folder, progress_callback=None):
    base_dir = os.path.dirname(os.path.abspath(__file__))
    procesados_dir = os.path.join(base_dir, "procesados")
    otros_dir = os.path.join(procesados_dir, "Otros")
    os.makedirs(procesados_dir, exist_ok=True)
    os.makedirs(otros_dir, exist_ok=True)

    descomprimir_archivos_en_carpeta(folder)
    archivos = obtener_todos_los_archivos(folder)
    total = len(archivos)

    facturas = []
    notas_credito = []
    mensajes_hacienda = []
    tiquetes = []
    pdfs = []
    otros = []
    mensajes_receptor = []

    def nombre_corto(nombre, maxlen=30):
        nombre = os.path.basename(nombre)
        return (nombre[:maxlen] + "...") if len(nombre) > maxlen else nombre

    for idx, archivo in enumerate(archivos):
        if progress_callback:
            progress_callback(idx + 1, total, nombre_corto(archivo))
        if archivo.lower().endswith(".xml"):
            limpiar_xml(archivo)
            try:
                tree = ET.parse(archivo)
                root = tree.getroot()
                # --- FACTURA ---
                if root.tag.endswith("FacturaElectronica"):
                    clave = None
                    fecha = None
                    for elem in root.iter():
                        if elem.tag.endswith("Clave"):
                            clave = elem.text
                        if elem.tag.endswith("FechaEmision"):
                            fecha = elem.text
                    if fecha:
                        mes, anio = obtener_mes_anio(fecha)
                    else:
                        mes, anio = obtener_mes_anio_desde_clave(clave)
                    facturas.append(
                        {"clave": clave, "path": archivo, "mes": mes, "anio": anio}
                    )
                # --- NOTA DE CRÉDITO ---
                elif root.tag.endswith("NotaCreditoElectronica"):
                    clave = None
                    cedula = None
                    referencia = None
                    fecha = None
                    for elem in root.iter():
                        if elem.tag.endswith("Clave"):
                            clave = elem.text
                        if elem.tag.endswith("Emisor"):
                            for subelem in elem.iter():
                                if subelem.tag.endswith("Numero"):
                                    cedula = subelem.text
                        if elem.tag.endswith("InformacionReferencia"):
                            for subelem in elem.iter():
                                if subelem.tag.endswith("Numero"):
                                    referencia = subelem.text
                        if elem.tag.endswith("FechaEmision"):
                            fecha = elem.text
                    if fecha:
                        mes, anio = obtener_mes_anio(fecha)
                    else:
                        mes, anio = obtener_mes_anio_desde_clave(clave)
                    notas_credito.append(
                        {
                            "clave": clave,
                            "path": archivo,
                            "cedula": cedula,
                            "referencia": referencia,
                            "mes": mes,
                            "anio": anio,
                        }
                    )
                # --- MENSAJE HACIENDA ---
                elif root.tag.endswith("MensajeHacienda"):
                    clave = None
                    cedula = None
                    fecha = None
                    for elem in root.iter():
                        if elem.tag.endswith("Clave"):
                            clave = elem.text
                        if elem.tag.endswith("NumeroCedulaReceptor"):
                            cedula = elem.text
                        if elem.tag.endswith("FechaEmision"):
                            fecha = elem.text
                    if fecha:
                        mes, anio = obtener_mes_anio(fecha)
                    else:
                        mes, anio = obtener_mes_anio_desde_clave(clave)
                    mensajes_hacienda.append(
                        {
                            "clave": clave,
                            "path": archivo,
                            "cedula": cedula,
                            "mes": mes,
                            "anio": anio,
                        }
                    )
                # --- TIQUETE ---
                elif root.tag.endswith("TiqueteElectronico"):
                    clave = None
                    fecha = None
                    for elem in root.iter():
                        if elem.tag.endswith("Clave"):
                            clave = elem.text
                        if elem.tag.endswith("FechaEmision"):
                            fecha = elem.text
                    if fecha:
                        mes, anio = obtener_mes_anio(fecha)
                    else:
                        mes, anio = obtener_mes_anio_desde_clave(clave)
                    tiquetes.append(
                        {"clave": clave, "path": archivo, "mes": mes, "anio": anio}
                    )
                # --- MENSAJE RECEPTOR ---
                elif root.tag.endswith("MensajeReceptor"):
                    clave = None
                    cedula = None
                    fecha = None
                    for elem in root.iter():
                        if elem.tag.endswith("Clave"):
                            clave = elem.text
                        if elem.tag.endswith("NumeroCedulaReceptor"):
                            cedula = elem.text
                        if elem.tag.endswith("FechaEmision"):
                            fecha = elem.text
                    if fecha:
                        mes, anio = obtener_mes_anio(fecha)
                    else:
                        mes, anio = obtener_mes_anio_desde_clave(clave)
                    mensajes_receptor.append(
                        {
                            "clave": clave,
                            "path": archivo,
                            "cedula": cedula,
                            "mes": mes,
                            "anio": anio,
                        }
                    )
                else:
                    otros.append(archivo)
            except Exception:
                otros.append(archivo)
        elif archivo.lower().endswith(".pdf"):
            try:
                texto = leer_pdf(archivo)
                import re

                texto_unido = texto.replace("\n", "").replace("\r", "")
                match = re.search(r"(?<!\d)(\d{50})(?!\d)", texto_unido)
                clave = match.group(1) if match else None
                factura = next((f for f in facturas if f["clave"] == clave), None)
                if factura:
                    mes = factura["mes"]
                    anio = factura["anio"]
                else:
                    mes, anio = obtener_mes_anio_desde_clave(clave)
                if clave:
                    pdfs.append(
                        {"clave": clave, "path": archivo, "mes": mes, "anio": anio}
                    )
                else:
                    otros.append(archivo)
            except Exception:
                otros.append(archivo)
        else:
            otros.append(archivo)

    # Organización por año y mes
    for factura in facturas:
        clave = factura["clave"]
        path = factura["path"]
        mes = factura["mes"]
        anio = factura["anio"]
        cedula = obtener_cedula_receptor(path)
        if not clave or not cedula:
            otros.append(path)
            continue
        cedula_dir = os.path.join(procesados_dir, cedula, anio, mes)
        facturas_dir = os.path.join(cedula_dir, "Facturas")
        os.makedirs(facturas_dir, exist_ok=True)
        destino = os.path.join(facturas_dir, f"FE-{clave}.xml")
        shutil.copy2(path, destino)

    for nc in notas_credito:
        clave = nc["clave"]
        path = nc["path"]
        mes = nc["mes"]
        anio = nc["anio"]
        cedula = obtener_cedula_receptor(path) or "Desconocido"
        if not clave or not cedula:
            otros.append(path)
            continue
        cedula_dir = os.path.join(procesados_dir, cedula, anio, mes)
        nc_dir = os.path.join(cedula_dir, "NotasCredito")
        os.makedirs(nc_dir, exist_ok=True)
        destino = os.path.join(nc_dir, f"NC-{clave}.xml")
        shutil.copy2(path, destino)

    for mh in mensajes_hacienda:
        clave = mh["clave"]
        path = mh["path"]
        mes = mh["mes"]
        anio = mh["anio"]
        cedula = obtener_cedula_receptor(path) or "Desconocido"
        if not clave or not cedula:
            otros.append(path)
            continue
        cedula_dir = os.path.join(procesados_dir, cedula, anio, mes)
        respuestas_dir = os.path.join(cedula_dir, "Respuestas")
        os.makedirs(respuestas_dir, exist_ok=True)
        destino = os.path.join(respuestas_dir, f"MH-{clave}.xml")
        shutil.copy2(path, destino)

    for mr in mensajes_receptor:
        clave = mr["clave"]
        path = mr["path"]
        mes = mr["mes"]
        anio = mr["anio"]
        cedula = obtener_cedula_receptor(path) or "Desconocido"
        if not clave or not cedula:
            otros.append(path)
            continue
        cedula_dir = os.path.join(procesados_dir, cedula, anio, mes)
        respuestas_dir = os.path.join(cedula_dir, "Respuestas")
        os.makedirs(respuestas_dir, exist_ok=True)
        destino = os.path.join(respuestas_dir, f"MR-{clave}.xml")
        shutil.copy2(path, destino)

    for tiquete in tiquetes:
        clave = tiquete["clave"]
        path = tiquete["path"]
        mes = tiquete["mes"]
        anio = tiquete["anio"]
        cedula = obtener_cedula_receptor(path)
        if not clave or not cedula:
            otros.append(path)
            continue
        cedula_dir = os.path.join(procesados_dir, cedula, anio, mes)
        tiquetes_dir = os.path.join(cedula_dir, "Tiquetes")
        os.makedirs(tiquetes_dir, exist_ok=True)
        destino = os.path.join(tiquetes_dir, f"TQ-{clave}.xml")
        shutil.copy2(path, destino)

    for pdf in pdfs:
        clave = pdf["clave"]
        path = pdf["path"]
        mes = pdf["mes"]
        anio = pdf["anio"]
        factura = next((f for f in facturas if f["clave"] == clave), None)
        cedula = None
        if factura:
            cedula = obtener_cedula_receptor(factura["path"])
        else:
            nc = next((n for n in notas_credito if n["clave"] == clave), None)
            cedula = obtener_cedula_receptor(nc["path"]) if nc else "Desconocido"
        if not cedula:
            cedula = "Desconocido"
        cedula_dir = os.path.join(procesados_dir, cedula, anio, mes)
        pdf_dir = os.path.join(cedula_dir, "PDF")
        os.makedirs(pdf_dir, exist_ok=True)
        destino = os.path.join(pdf_dir, f"PDF-{clave}.pdf")
        shutil.copy2(path, destino)

    for archivo in otros:
        ext = os.path.splitext(archivo)[1].lower().replace(".", "") or "sin_extension"
        ext_dir = os.path.join(otros_dir, ext)
        os.makedirs(ext_dir, exist_ok=True)
        try:
            shutil.copy2(archivo, os.path.join(ext_dir, os.path.basename(archivo)))
        except Exception:
            pass

    generar_reportes_excel(facturas, tiquetes, notas_credito)
